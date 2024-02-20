# pip install openpyxl

import openpyxl
import Utils
from openpyxl.styles import Font, Fill, Border, Alignment
from copy import copy


class EmployeeICHRAInfo:
    def __init__(self, ssn, dob, home_zip, work_zip, salary, lcsp):
        self.ssn = ssn
        self.dob = dob
        self.home_zip = home_zip
        self.work_zip = work_zip
        self.salary = salary
        self.lcsp = lcsp

    def __str__(self):
        return f"SSN: {self.ssn}, Salary: {self.salary}, DOB: {self.dob}, Work Zip: {self.work_zip}, Home Zip: {self.home_zip}, LCSP: {self.lcsp}"


def copy_cell(source_cell, target_cell):
    # Copy value
    target_cell.value = source_cell.value

    # Copy style attributes
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)


def get_employee_map_from_file(file_path, zip_to_use):
    Utils.print_line_separator()
    employees = {}

    sheet = get_active_sheet(file_path)

    for row in sheet.iter_rows(min_row=5):
        missing_dob, missing_zip = is_dob_and_zip_are_valid(row, zip_to_use)

        if missing_zip and not missing_dob:
            print("Zip is missing for employee with SSN: " + row[2].value)
        elif missing_dob and not missing_zip:
            print("Date of birth is missing for employee with SSN: " + row[2].value)
        elif missing_dob and missing_zip:
            print("Date of birth and zip are missing for employee with SSN: " + row[2].value)

        ssn = row[2].value
        dob = row[22].value
        home_zip = row[27].value
        work_zip = row[31].value
        salary = row[34].value
        lcsp = row[38].value
        lcsp = None if lcsp == "" else lcsp

        employee = EmployeeICHRAInfo(ssn, dob, home_zip, work_zip, salary, lcsp)

        employees[employee.ssn] = employee
    return employees


def is_dob_and_zip_are_valid(row, zip_to_use):
    missing_zip = False
    missing_dob = False
    if row[2].value is None:
        print("SSN is missing for a row number: " + str(row[0].row))
    if row[22].value is None or row[22].value == "N/A" or row[22].value == "None" or row[22].value == "":
        missing_dob = True
    if zip_to_use == "Home Zip" and (
            row[27].value is None or row[27].value == "N/A" or row[27].value == "None" or row[27].value == ""):
        missing_zip = True
    if zip_to_use == "Work Zip" and (
            row[31].value is None or row[31].value == "N/A" or row[31].value == "None" or row[31].value == ""):
        missing_zip = True
    return missing_dob, missing_zip


def write_employee_map_to_file(file_path, employees, zip_to_use):
    original_wb = openpyxl.load_workbook(file_path)

    # Create a new workbook
    new_wb = openpyxl.Workbook()
    # Remove the default sheet
    new_wb.remove(new_wb.active)

    # Iterate through each sheet in the original workbook
    for sheet_name in original_wb.sheetnames:
        original_ws = original_wb[sheet_name]

        # Create a new sheet in the new workbook with the same name
        new_ws = new_wb.create_sheet(title=sheet_name)

        # Copy data and style from each cell of the original sheet
        for row in original_ws.iter_rows():
            for cell in row:
                new_cell = new_ws.cell(row=cell.row, column=cell.column)
                copy_cell(cell, new_cell)

    new_wb._active_sheet_index = 0
    sheet = new_wb.active

    for row in sheet.iter_rows(min_row=5):
        ssn = row[2].value
        for cell in row:
            if cell.column == 39 and ssn in employees:
                employee = employees[ssn]
                if employee.lcsp is not None:
                    cell.value = str(employee.lcsp).replace('$', '')

    zip_to_use = zip_to_use.replace(" ", "_").lower()

    # if filepath has _with_LCSP_by_ in it do not change the name
    if '_with_LCSP_by_' not in file_path:
        output_file = file_path.replace('.xlsx', '_with_LCSP_by_' + zip_to_use + '.xlsx')
    else:
        output_file = file_path

    print("Saving the file...")

    new_wb.save(output_file)

    print("File saved: \n" + output_file)


def get_active_sheet(file_path):
    file = openpyxl.load_workbook(file_path)
    file._active_sheet_index = 0
    sheet = file.active
    return sheet
